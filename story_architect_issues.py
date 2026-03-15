from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Issue Tracker"

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='1F4E79')
BODY_FONT = Font(name='Arial', size=10)
BODY_FONT_BOLD = Font(name='Arial', size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
WRAP_CENTER = Alignment(wrap_text=True, vertical='top', horizontal='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

# Severity colors
SEV_COLORS = {
    'Critical': PatternFill('solid', fgColor='FFCCCC'),
    'High': PatternFill('solid', fgColor='FFE0CC'),
    'Medium': PatternFill('solid', fgColor='FFFFCC'),
    'Low': PatternFill('solid', fgColor='E0FFE0'),
}

# Column widths
COL_WIDTHS = {
    'A': 5,   # #
    'B': 8,   # Priority
    'C': 35,  # Issue
    'D': 12,  # Severity
    'E': 12,  # Category
    'F': 14,  # Status
    'G': 50,  # Recommended Fix
    'H': 45,  # Potential Problems from Fix
    'I': 14,  # Effort
    'J': 45,  # Notes / Additional Context
}

for col, width in COL_WIDTHS.items():
    ws.column_dimensions[col].width = width

# Headers
headers = ['#', 'Priority', 'Issue', 'Severity', 'Category', 'Current Status', 'Recommended Fix', 'Potential Problems from Fix', 'Effort', 'Notes / Additional Context']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = WRAP_CENTER
    cell.border = THIN_BORDER

ws.row_dimensions[1].height = 30

# Data rows — ordered by recommended implementation priority
issues = [
    # --- PHASE 1: Quick wins / critical fixes ---
    ['PHASE 1: QUICK WINS & CRITICAL FIXES', None],

    [1, 'P1', 'Feature flag was not active for prior sessions',
     'Critical', 'Config', 'Resolved — flag now set',
     'Verify by running a test session and checking for cultural brief data in the session JSON. Add ENABLE_CULTURAL_ENGINE=true to any .env.example or deployment docs so it doesn\'t get missed again.',
     'None — this is already fixed. Risk is only if deployed without the env var.',
     'Trivial',
     'The flag is set in .env now. Any session created before the flag was added will have no cultural data. Consider removing the feature flag entirely once the engine is validated, since it should always be on.'],

    [2, 'P1', 'Cultural engine never tested end-to-end in a real session',
     'Critical', 'Testing', 'Unknown — needs verification',
     'Run a test session with a culturally rich seed (e.g., your "gay twink" premise). After clarifier turn 2, inspect the session JSON for cultural brief data. Check the server logs for [CULTURAL] prefixed messages. Verify the brief text appears in the prompt history.',
     'If the engine produces bad briefs, they could actively harm output quality. Need a kill switch (the feature flag) and a quality check before going wide.',
     'Low',
     'The two-LLM-call architecture (summarizer + researcher) on Haiku may produce thin briefs. May need to bump to Sonnet for the researcher role if brief quality is insufficient. Test with 2-3 different cultural contexts.'],

    [3, 'P1', 'sceneService has zero cultural engine integration',
     'High', 'Code Gap', 'Confirmed — no cultural imports',
     'Add the same integration pattern used in hookService/characterService: import culturalResearchService, add getCulturalBrief/getCulturalBriefForBuilder/fireBackgroundCulturalResearch helpers, inject brief into buildClarifierPrompt and buildBuilderPrompt, fire background research after clarifier turns.',
     'Increases scene generation token cost (2 extra Haiku calls per scene if throttle fires). Could slow down an already-slow module. Scene prompts are already very long — adding cultural context may push against context limits.',
     'Medium',
     'Scenes are the final creative output — cultural grounding matters most here (dialogue authenticity, setting details, culturally specific behavior). But the scene module is already the slowest, so cultural calls should be aggressively throttled. Consider only firing for scenes flagged as culturally relevant.'],

    [4, 'P2', 'Divergence engine and cultural engine outputs invisible to user',
     'High', 'UX Gap', 'Confirmed — no UI surface',
     'Add a "Debug/Insights" panel accessible via a button on each module\'s clarifier screen. Show: (1) cultural brief with evidence items and creative applications, (2) divergence direction map with alternative paths, (3) psychology signals and adaptation plan. Use a collapsible drawer or modal — not inline, to avoid cluttering the main flow.',
     'Users may over-rely on debug info and second-guess the engine. Could create "backseat driver" behavior where they try to manually steer the cultural/divergence outputs instead of trusting the system. Also adds frontend complexity.',
     'Medium-High',
     'For testing purposes this is essential. Consider making it a "developer mode" toggle that shows engine internals. Could also log these to a separate "engine insights" tab per session for post-hoc analysis.'],

    # --- PHASE 2: Story quality & user control ---
    ['PHASE 2: STORY QUALITY & USER CONTROL', None],

    [5, 'P2', 'Story length always roughly the same (12-20 beats hardcoded)',
     'High', 'Prompt/Architecture', 'Confirmed — beat count doesn\'t scale with scope',
     'Make the plot builder\'s beat count guidance dynamic. Read the scope/length constraint from the hook\'s constraint ledger. Map: "2-hour gut-punch" -> 6-10 beats, "standard" -> 12-18 beats, "slow-burn season" -> 20-35 beats, "episodic series" -> 30-50+ beats (chunked). Pass the mapped range into the plot builder prompt template as a variable.',
     'Very long stories (30+ beats) may exceed context window limits for the builder LLM. Beat quality tends to degrade at high counts — later beats become generic. Scene module would need to handle 20+ scenes, compounding the "takes too long" issue. May need a chunked building approach for long stories.',
     'Medium',
     'The scope/length signal already exists in the hook constraint ledger — it just doesn\'t flow into the beat count. This is a relatively clean fix. For very long stories, consider a "chapter" abstraction between plot and scene that groups beats into narrative arcs.'],

    [6, 'P2', 'No "loose ends resolver" phase before scene generation',
     'High', 'Architecture', 'Confirmed — development targets exist but are invisible',
     'Add a "Pre-Scene Audit" step between plot lock and scene planning. This step: (1) aggregates all unresolved development targets from hook/character/world/plot judges, (2) presents them to the user as a checklist with severity and suggested fixes, (3) lets the user choose to address each (with options) or accept as-is, (4) feeds resolved decisions back into the scene planning phase.',
     'Adds another interaction step to an already long pipeline. Users who want fast generation will find it annoying. Could create a "too many problems" feeling that makes users lose confidence in earlier modules. Some issues may not be fixable at this stage.',
     'High',
     'The upstream development target system is already robust — judges produce weaknesses, they\'re carried through packs, downstream modules get prompted to address them. The gap is purely that the user never sees this. The audit step surfaces existing data. Consider making it skippable ("Everything looks good — proceed?" vs "We found 3 things worth reviewing").'],

    [7, 'P2', 'Characters default to ~4 regardless of story needs',
     'Medium', 'Prompt/LLM', 'Partially addressed — prompt says 4-6 but LLM anchors at 4',
     'Three changes: (1) Surface cast size as an explicit assumption in the character clarifier\'s first turn with vivid alternatives ("A tight 3-person pressure cooker" / "A 5-person crew where alliances shift" / "A 7-character court with factions"). (2) Tie cast size to scope/length — a "slow-burn season" needs more characters. (3) In the builder prompt, add: "The user chose [N] characters. Generate exactly [N] character profiles."',
     'Larger casts increase token cost, builder complexity, and scene count. Characters 5+ tend to get less psychological depth. The character clarifier would need more turns for larger casts. Supporting characters in large ensembles risk becoming "furniture" the prompt already warns about.',
     'Medium',
     'The prompt already says "Most stories need 4-6" which anchors the LLM. Changing to an explicit user-facing assumption gives the user direct control. The real challenge is maintaining character quality at scale — each additional character needs its own want/misbelief/stakes, which means more clarifier turns or more aggressive inference.'],

    [8, 'P2', 'No opportunity to edit character details (race, age, name) before builder',
     'High', 'UX Gap', 'Confirmed — no pre-builder review step',
     'Add a "Character Review" screen after the clarifier reaches readiness but before the builder fires. Show each character\'s resolved state (role, inferred traits, confirmed traits) with editable fields for: name, age, race/ethnicity, gender, physical description, and any other concrete details. The system provides recommendations (inferred from the story context) but the user can override any field. Feed the overrides into the builder prompt.',
     'Adds a mandatory interaction step. Could break the "fun" flow if it feels like filling out a form. Users may make choices that conflict with the story logic (e.g., setting a character\'s age at 8 in an adult romance). Need validation or at least a warning system. Also, some details (like name) may need to be consistent with the world/setting — a medieval Japanese setting shouldn\'t have a character named "Brad."',
     'Medium-High',
     'This is one of the most user-requested features based on your feedback. The key is making it feel like a fun creative moment, not a form. Show it as "Meet your cast — any last changes before they come to life?" with smart defaults. Consider tying this to the cultural engine — it could suggest culturally appropriate names and physical descriptions.'],

    [9, 'P2', 'No warning when user steers story in a bad direction',
     'Medium', 'Prompt/UX', 'Partially addressed — conflict_flag exists but limited',
     'Strengthen the conflict_flag system: (1) Add a "story health score" that aggregates flags across turns. (2) When health drops below a threshold, show a prominent warning with specific problems and fix options — not just a flag. (3) Add a "quality gate" before each builder phase where accumulated flags must be resolved or explicitly accepted. (4) Give the user 2-3 fix options for each flag, not just a description of the problem.',
     'Over-warning kills creative freedom — users will feel nannied. The "quality gate" could block users who have a valid creative vision the system doesn\'t understand. Edge cases and experimental stories will trigger false positives. Need a "I know what I\'m doing, proceed anyway" escape hatch.',
     'Medium',
     'The existing conflict_flag field is the right mechanism — it just needs to be more prominent in the UI and offer actionable options instead of just descriptions. The prompt already says "The UI will show this as a warning with options to fix it" but the options need to be concrete choices, not just "fix this."'],

    # --- PHASE 3: Performance & readability ---
    ['PHASE 3: PERFORMANCE & READABILITY', None],

    [10, 'P3', 'Scene module takes too long / user input often unnecessary',
     'High', 'Performance/UX', 'Partially addressed — auto-pass exists but undertriggered',
     'Three-pronged fix: (1) Add a "Generate All" mode where all scenes build without interaction (user reviews at the end). (2) Increase auto-pass aggressiveness — lower the confidence threshold from 0.85 to 0.70 for non-turning-point scenes. (3) Batch-build scenes in parallel (3-4 at a time) instead of sequentially. (4) Skip the minor judge for all but turning-point and climax scenes.',
     'Parallel building increases API cost burst. "Generate All" mode means user has no mid-stream steering — if scene 3 goes wrong, scenes 4-20 build on a bad foundation. Aggressive auto-pass may skip scenes that actually need user input. Skipping judges means quality issues go undetected until the user reads the final output.',
     'High',
     'The scene module is the biggest time sink because it\'s N scenes x (clarifier + builder + optional judge) calls. For a 15-scene story, that\'s potentially 45+ LLM calls. The "Generate All" mode is the biggest win — most users want to see the final story, not micromanage each scene. Add a "review and revise" step after bulk generation.'],

    [11, 'P3', 'Scene output hard to read — should be more script-like',
     'Medium', 'Frontend', 'Likely frontend rendering issue — data model is already script-like',
     'The scene builder already produces screenplay-format readable_scene with character names in caps, parenthetical delivery notes, stage directions as prose blocks, and internal monologue in italics. Fix is frontend: (1) Render the screenplay_text field with proper formatting — monospace or screenplay font, character names bolded/capped, stage directions in gray italics, dialogue indented. (2) Add a "Script View" toggle that shows clean screenplay format vs raw data.',
     'Custom screenplay rendering adds frontend complexity. Different users may want different formats (novel-style vs screenplay vs VN-preview). Multiple view modes means multiple rendering paths to maintain.',
     'Medium',
     'Check if the frontend is displaying readable_scene.screenplay_text or if it\'s showing raw JSON/vn_scene data. The fix may be as simple as rendering the existing screenplay_text field with proper CSS. Consider also adding a "VN Preview" mode that shows the scene as it would appear in a visual novel (with character portraits and text boxes).'],

    [12, 'P3', 'Some modules can\'t show preview when loading exports',
     'Medium', 'Frontend', 'Likely — needs frontend investigation',
     'Audit the frontend pack preview components for each module. Each pack type (HookPack, CharacterPack, WorldPack, PlotPack, ScenePack, CharacterImagePack) needs a dedicated preview renderer that displays the key locked fields in a readable format. Add missing renderers and ensure the preview route/component handles all pack types.',
     'Each pack has a different structure — building 6 custom preview renderers is non-trivial. Pack structures may evolve, requiring preview component updates. Large packs (especially plot with 20+ beats) need summarization or pagination.',
     'Medium',
     'This is a frontend-only fix. The data exists — it\'s just not being displayed. Priority should be: hook (most commonly loaded), character, plot, world, scene, characterImage. The state_summary field in each pack is designed for exactly this purpose.'],

    # --- PHASE 4: LLM behavior & consistency ---
    ['PHASE 4: LLM BEHAVIOR & CONSISTENCY', None],

    [13, 'P3', 'Prior user statements not carried forward / implications ignored',
     'High', 'Prompt/LLM', 'Architecture is correct — LLM execution inconsistent',
     'Three reinforcements: (1) Move the constraint ledger to the END of the user prompt (LLMs attend more to the end). (2) Add a "MUST HONOR" section that lists only confirmed constraints in a compact, high-signal format. (3) In the builder prompt, repeat confirmed constraints as hard requirements, not context. (4) Consider a lightweight "constraint checker" LLM call that verifies builder output against confirmed constraints before accepting it.',
     'Moving ledger position may displace other important context. A constraint checker adds another LLM call per build. Over-constraining the builder may produce rigid, formulaic output. If the "MUST HONOR" list is long, it becomes noise.',
     'Medium',
     'This is fundamentally an LLM attention problem. The constraint ledger is already in the prompt — the LLM just doesn\'t always obey it, especially in long contexts. The highest-impact fix is probably the compact "MUST HONOR" block at prompt end. Also consider: when the user explicitly states something, mark it as "user_direct_statement" source in the ledger (higher priority than "user_chose" from a chip).'],

    [14, 'P3', 'Characters too similar to each other',
     'Medium', 'Prompt/LLM', 'Partially addressed — diversity scoring exists in judge',
     'Four improvements: (1) Add a "differentiation matrix" to the builder prompt requiring each character to differ on at least 3 of: stress response, communication style, core value, fear type, power strategy. (2) In the judge, make diversity a hard-fail criterion (not just a score). (3) Add explicit contrast pairs to the builder output: "Character A handles conflict by X, Character B by Y — they\'re designed to clash." (4) Surface character comparison in the pre-builder review so the user can see if characters feel too similar.',
     'Forcing differentiation can lead to artificial/forced character traits. The LLM may default to simple trait inversions (brave/cowardly, loud/quiet) rather than meaningful differentiation. Hard-failing on diversity may cause excessive rebuild loops.',
     'Medium',
     'The builder already has a structural_diversity field with a boolean and explanation. The judge scores diversity. But neither is aggressive enough about enforcement. The differentiation matrix is the key addition — it gives the LLM a concrete checklist for what "different" means.'],

    [15, 'P3', 'Assumptions may not be deterministic — different runs give different assumptions',
     'Low', 'Design', 'By design — LLM generates assumptions dynamically each turn',
     'If reproducibility is desired: (1) Add a "core assumptions" template per story archetype that always gets surfaced on turn 1 (setting, tone, protagonist type, relationship type, scope). (2) Let the LLM add 2-3 additional dynamic assumptions on top. (3) Consider using a fixed LLM temperature (e.g., 0.3) for assumption generation to reduce variance. Current temp is likely default (~1.0).',
     'Fixed assumption templates make the experience feel scripted and repetitive across sessions. Lower temperature reduces creative surprise. Template assumptions may not fit all seed types (a detailed seed needs different assumptions than a vague one).',
     'Low',
     'The current dynamic approach is actually good — it means each session feels fresh. The potential issue is when the SAME seed produces wildly different assumption sets on different runs, which could confuse users. A middle ground: ensure the same seed always surfaces the SAME categories of assumptions (setting, tone, protagonist) but let the specific content vary.'],

    [16, 'P4', 'Drift concerns as user progresses through modules',
     'Medium', 'Architecture', 'Partially addressed — constraint ledger + psych ledger carry forward',
     'Add a "Story Bible" document that\'s built incrementally as each module locks. The bible contains: (1) all confirmed constraints from all modules, (2) locked creative decisions with their sources, (3) character profiles as they evolve, (4) world rules, (5) plot beats. Each new module receives the full bible as a compact reference document. Add a "consistency check" at each module boundary that flags contradictions between the new module\'s output and the bible.',
     'The story bible grows with each module — could become very large and hit context limits. Building the bible requires a summarization step. The consistency check adds another LLM call per module transition. May create rigidity — some "drift" is actually creative evolution.',
     'High',
     'The packs already serve as a partial story bible, but they\'re structured data, not a readable narrative summary. The key insight is that LLMs work better with a coherent prose summary than with structured JSON. A 500-word "story bible" prose summary regenerated at each module boundary would be more effective than the current pack-passing approach. This is the highest-effort fix but addresses drift most comprehensively.'],

    [17, 'P3', 'Unclear if cross-module problem fixing is actually working',
     'Medium', 'Transparency', 'Working as designed — but invisible to user',
     'Surface the upstream development targets system: (1) In the debug/insights panel (Issue 4), show active development targets with their status. (2) When a downstream module addresses a target, show a subtle "Improved: [area]" indicator. (3) In the pre-scene audit (Issue 6), show the full target history with which modules addressed which gaps.',
     'Revealing the "fixing" system may make users feel the earlier modules were broken. Could undermine confidence. Showing too much internal machinery breaks the "fun creative partner" illusion.',
     'Low (if Issues 4 and 6 are built)',
     'This is essentially free once you build the debug panel (Issue 4) and pre-scene audit (Issue 6). The data already exists in the pack handoffs. Just needs rendering.'],

    # --- PHASE 5: Additional issues discovered during review ---
    ['PHASE 5: ADDITIONAL ISSUES DISCOVERED DURING REVIEW', None],

    [18, 'P2', 'Hook module has no explicit "character count" assumption — cast size never directly discussed with user',
     'Medium', 'Prompt Gap', 'Not addressed — cast size only mentioned in character prompt guidance',
     'Add a "CAST SIZE" creative constraint to the hook clarifier (alongside scope/length, genre feel, etc.) that gets surfaced as an assumption when relevant. Map to specific ranges: "intimate duo" (2-3), "pressure triangle" (3-4), "ensemble crew" (5-7), "faction epic" (7+). Carry the confirmed cast size into the character module as an imported constraint.',
     'Asking about cast size in the hook module may be premature — users may not know how many characters they want until the story takes shape. Could feel like a form question if not framed as a vivid creative choice.',
     'Low',
     'This directly feeds Issue 7 (characters defaulting to 4). If cast size is confirmed in hook and imported into character, the character module has a concrete target. Frame it as a story-feel choice, not a number: "Is this a locked-room two-person showdown or a Game-of-Thrones faction war?"'],

    [19, 'P3', 'Psychology ledger may accumulate stale/contradictory signals over 6 modules',
     'Medium', 'Architecture', 'Partially addressed — consolidation exists but may not be aggressive enough',
     'Add a "psychology reset" at each module boundary that: (1) runs the consolidation function, (2) drops signals with confidence < 0.2, (3) merges duplicate signals, (4) caps total active signals at 8-10 (keeps highest confidence). This prevents the ledger from becoming noise.',
     'Aggressive pruning may discard signals that become relevant later. Resetting at module boundaries loses nuance. The confidence threshold may be too aggressive for slow-building patterns.',
     'Low',
     'The consolidation system already exists (shouldConsolidate throttle, runConsolidation, applyConsolidation) but fires based on throttle timing, not module boundaries. A forced consolidation at lock time would be a clean addition. Check if the plot module\'s "MODULE 5 PSYCHOLOGY SHIFT" guidance (line 160-165 in plotPrompts.ts) is being followed — it says to stop re-confirming old signals.'],

    [20, 'P4', 'No way to go back and revise a locked module without restarting',
     'Medium', 'UX Gap', 'Not addressed — modules are strictly linear',
     'Add an "Unlock and Revise" feature that lets the user go back to a previously locked module. The revised module re-runs its builder and judge, then all downstream modules are marked as needing regeneration. Show a warning: "Changing [module] will require regenerating [downstream modules]."',
     'Extremely complex architecturally. Downstream modules may have made creative decisions based on the original output. Regenerating cascades could produce very different results. Users may lose work they liked in downstream modules. The session state management becomes much more complex.',
     'Very High',
     'This is a major feature that many creative tools support. For now, the workaround is starting a new session from the same seed. A lighter version: allow re-running just the builder/judge of the current module (before proceeding to the next one) with modified constraints.'],

    [21, 'P3', 'Token cost and latency not tracked or displayed to user',
     'Low', 'Observability', 'Not addressed — no cost/performance tracking visible',
     'Add a per-session and per-module token usage tracker. Display: (1) total tokens used, (2) estimated cost, (3) time per LLM call, (4) number of rebuild loops (judge failures). Show in the debug panel. This helps with optimization and gives users transparency about what the system is doing.',
     'Cost display may alarm users. Showing "this cost $2.50 in API calls" could make them hesitant to iterate. Time tracking adds overhead.',
     'Low',
     'Useful for development and optimization. The LLMClient probably already has response metadata (tokens used, latency). Just needs aggregation and display. Could also help identify which modules/prompts are most expensive and need optimization.'],

    [22, 'P2', 'Builder tournament/retry loop has no user visibility or cap',
     'Medium', 'UX/Architecture', 'Partially addressed — tournament exists but invisible',
     'Surface the build-judge loop to the user: (1) Show "Building... attempt 1/3" with a brief reason if the judge fails ("specificity too low — regenerating"). (2) Cap retries at 3 with a "best effort" fallback. (3) After max retries, show the user what the judge flagged and let them choose: accept as-is, retry with adjusted prompt, or go back to clarifier.',
     'Showing failures may reduce confidence. The "best effort" fallback may produce below-threshold output. Letting users adjust prompts is powerful but complex.',
     'Medium',
     'The tournament/retry system is already built (TournamentProgress type exists in hook.ts). Users currently see a loading spinner with no idea what\'s happening. Even just "Building your story... quality checking..." would help. The 3-retry cap with user fallback is the key safety valve.'],
]

# Write data
row = 2
for item in issues:
    if item[1] is None:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=item[0])
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(vertical='center')
        cell.border = THIN_BORDER
        for c in range(2, 11):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.row_dimensions[row].height = 28
        row += 1
        continue

    for c, val in enumerate(item, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    # Severity coloring
    sev = item[3]
    if sev in SEV_COLORS:
        ws.cell(row=row, column=4).fill = SEV_COLORS[sev]

    # Center-align # and priority
    ws.cell(row=row, column=1).alignment = WRAP_CENTER
    ws.cell(row=row, column=2).alignment = WRAP_CENTER
    ws.cell(row=row, column=4).alignment = WRAP_CENTER
    ws.cell(row=row, column=5).alignment = WRAP_CENTER
    ws.cell(row=row, column=6).alignment = WRAP_CENTER
    ws.cell(row=row, column=9).alignment = WRAP_CENTER

    ws.row_dimensions[row].height = 100
    row += 1

# Freeze panes
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:J{row-1}'

# --- Summary sheet ---
ws2 = wb.create_sheet('Summary')
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 50

summary_headers = ['Metric', 'Count', 'Details']
for c, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = WRAP_CENTER
    cell.border = THIN_BORDER

summary_data = [
    ['Total Issues', 22, 'Including 5 additional issues discovered during review'],
    ['Critical Severity', 2, 'Feature flag verification + end-to-end testing'],
    ['High Severity', 7, 'Scene cultural integration, debug panel, length scaling, loose ends, character editing, scene speed, carry-forward'],
    ['Medium Severity', 10, 'Cast size, warnings, similarity, readability, previews, drift, psych signals, tournament, cast count, cross-module'],
    ['Low Severity', 3, 'Assumption determinism, token tracking, psych stale signals'],
    ['', '', ''],
    ['Category Breakdown', '', ''],
    ['Code Gap', 1, 'sceneService cultural integration'],
    ['Prompt/LLM Behavior', 4, 'Carry-forward, cast size, similarity, assumptions'],
    ['Architecture', 4, 'Length scaling, loose ends, drift/story bible, module revision'],
    ['UX Gap', 4, 'Debug panel, character editing, warnings, tournament visibility'],
    ['Frontend', 2, 'Scene readability, export previews'],
    ['Config/Testing', 2, 'Feature flag, end-to-end test'],
    ['Prompt Gap', 1, 'Cast count assumption in hook'],
    ['Performance/UX', 1, 'Scene module speed'],
    ['Transparency', 1, 'Cross-module fixing visibility'],
    ['Observability', 1, 'Token/cost tracking'],
    ['Design', 1, 'Assumption determinism'],
    ['', '', ''],
    ['Effort Breakdown', '', ''],
    ['Trivial', 1, 'Feature flag fix'],
    ['Low', 5, 'E2E test, psych reset, cast assumption, token tracking, cross-module visibility'],
    ['Medium', 7, 'Cultural integration, length scaling, cast defaults, warnings, similarity, frontend fixes'],
    ['Medium-High', 2, 'Debug panel, character editing screen'],
    ['High', 3, 'Loose ends resolver, scene speed improvements, story bible'],
    ['Very High', 1, 'Module unlock/revision feature'],
]

for r, row_data in enumerate(summary_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT if not (val == '' and c == 1) else BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if r in [2, 8, 21] or (isinstance(val, str) and val.endswith('Breakdown')):
            cell.font = BODY_FONT_BOLD

output_path = '/sessions/busy-awesome-hopper/mnt/visnovgen/story_architect_issues.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
